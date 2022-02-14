# encoding = utf-8
from itertools import product
from openpyxl import Workbook, load_workbook
from domain.caseitem import caseitem
import json
import gc, os
from typing import List
from domain.useritem import useritem
from domain.loginDTO import logindto
from module.passport.mapper import *


def loaduser(env: str) -> List[useritem]:
    """
    加载测试用户的数据
    :return: <List>[useritem]
    """
    userlist = []
    file = "userdata/{}/测试用户.xlsx".format(env)
    wb = load_workbook(file, data_only=True)
    ws = wb.worksheets[0]
    merged_cells = ws.merged_cells
    # print(merged_cells)
    for c in merged_cells:
        minrow = c.min_row
        # print(ws.cell(minrow, 3).value)
        userlist.append(useritem(accountinfo=json.loads(ws.cell(minrow, 3).value),
                                 app=ws.cell(minrow + 1, 3).value,
                                 client="",
                                 platform=ws.cell(minrow + 2, 3).value,
                                 userType=ws.cell(minrow + 3, 3).value,
                                 password=ws.cell(minrow + 4, 3).value,
                                 twoFa=ws.cell(minrow + 5, 3).value,
                                 disable=ws.cell(minrow + 6, 3).value,
                                 inblacklist=ws.cell(minrow + 7, 3).value))
    del ws, wb
    gc.collect()
    return userlist


def chooseuser(userlist: List[useritem], app, platform, accountType, userType, changjing) -> logindto:
    """
    筛选用户
    :param userlist: <List[useritem]> 所用测试数据，测试用户
    :param app: <str> app名称：['WEB', '生学堂', '云平台', '晓我课堂', '晓我课外', '晓我课堂平板', '壹教云']
    :param platform: <str> 平台名称： ['PC', '手机', '微信', '龙泉', '壹教云']
    :param accountType: <str> 登录方式：['手机号', '身份证号', '用户名', '手机验证码', '龙泉登录', '生学号']
    :param userType: <str> 用户类型，和client联动，只需要输入usertype，client即可匹配上：['学生', '教师', '家长', '管理员']
    :return: <logindto> 返回第一个满足条件的测试数据，测试用户
    """
    for u in userlist:
        # if accountType == u.accountType and userType in u.accountType and app in u.app and platform in u.platform:
        # print(accountType in list(u.accountinfo))
        # print(userType in u.userType)
        # print(app in u.app)
        # print(platform in u.platform)
        # print(u.accountinfo)
        # print(u.show())

        if accountType in list(u.accountinfo) and userType in u.userType and app in u.app and platform in u.platform:
            # print(u.accountinfo[accountType])
            if u.accountinfo[accountType] == '13550000726' and changjing == "已禁用的用户登录":
                pass
                # PC - 云平台 - 手机号 - 教师 - 已禁用的用户登录
            if changjing == "已禁用的用户登录":
                if u.disable == True:
                    return logindto(user=u,
                                    account=u.accountinfo[accountType],
                                    accountType=accountType_mapping[accountType],
                                    app=app_mapping[app],
                                    platform=platform_mapping[platform],
                                    userType=userType_mapping[userType],
                                    client=client_mapping[userType])
                else:
                    continue
            else:
                if u.disable == True:
                    continue

            if changjing == "黑名单":
                if u.inblacklist == True:
                    return logindto(user=u,
                                    account=u.accountinfo[accountType],
                                    accountType=accountType_mapping[accountType],
                                    app=app_mapping[app],
                                    platform=platform_mapping[platform],
                                    userType=userType_mapping[userType],
                                    client=client_mapping[userType])
                else:
                    continue
            else:
                if u.inblacklist == True:
                    continue

            return logindto(user=u,
                            account=u.accountinfo[accountType],
                            accountType=accountType_mapping[accountType],
                            app=app_mapping[app],
                            platform=platform_mapping[platform],
                            userType=userType_mapping[userType],
                            client=client_mapping[userType])
    return None


def bulidcontent(logindto: logindto, changjing):
    """
    生成请求传参，断言
    :param logindto: <logindto> 登录DTO
    :param changjing: <str> 测试场景：['正常', '错误密码', '账号为空', '密码为空', '账号过长', '密码过长', '已禁用的用户登录', 'XSS', "黑名单"]
    :return: <dict>,<str> 生成的请求参数
    """

    # 判断该DOT是否是使用手机验证码登录，如果是，则将默认密码替换成999999的加密密码 0Cic12Rv+wpC/3HY1md65A==
    if logindto.accountType == accountType_mapping["手机验证码"]:
        logindto.password = "0Cic12Rv+wpC/3HY1md65A=="
    if logindto is None: return None, None
    # 正常
    if changjing == changjing_list[0]:
        return logindto.ToDict(), '''"code" : 200'''
    # 错误密码
    if changjing == changjing_list[1]:
        logindto.password = wrong_passport
        return logindto.ToDict(), '''"code":8072401'''
    # 账号为空
    if changjing == changjing_list[2]:
        logindto.account = ""
        return logindto.ToDict(), '''"code":100101'''
    # 密码为空
    if changjing == changjing_list[3]:
        logindto.password = ""
        return logindto.ToDict(), '''"message":"密码不能为空"'''
    # 账号过长
    if changjing == changjing_list[4]:
        logindto.account = long_account
        return logindto.ToDict(), '''"message":"账号长度最小为 5，最大为 32 位"'''
    # 密码过长
    if changjing == changjing_list[5]:
        logindto.account = long_password
        return logindto.ToDict(), '''"message":"账号长度最小为 5，最大为 32 位"'''
    # 已禁用的用户登录
    if changjing == changjing_list[6]:
        return logindto.ToDict(), '''"message":"该账号已被禁用"'''
    # XSS
    if changjing == changjing_list[7]:
        logindto.account = XSS_text
        return logindto.ToDict(), '''"message":"账号只能由英文数字和下划线、中划线组成"'''
    # 用户被加入黑名单
    if changjing == changjing_list[8]:
        return logindto.ToDict(), '''"code":8072424'''
    return None, None


def build(env: str):
    """
    过滤并创建测试用例，包括："模块", "ID", "用例名称", "请求类型", "URL", "参数", "断言", "是否执行"
    :param userlist: 所有的测试用户数据
    """
    userlist: List[useritem] = loaduser(env)
    loopval = [list(platform_mapping), list(app_mapping), list(accountType_mapping), list(userType_mapping),
               changjing_list]
    excel_text = []
    excel_text.append(["模块", "ID", "用例名称", "请求类型", "URL", "参数", "断言", "是否执行"])
    for i in product(*loopval):
        excel_text.append(i)

    wb = Workbook()  # 打开excel
    ws = wb.active  # 打开当前活跃的sheet
    ws.title = "登录用例"
    ws.append(excel_text[0])  # 加入标题

    # 生成用例
    index_id = 1  # id起始
    for i in excel_text[1:]:
        casename = ""
        line = []
        for j in i:
            casename = casename + j + "-"
        # 这里添加业务过滤
        if filtercasename(casename):
            continue
        # 根据用例名称，结合录入的用户生成参数
        mo = casename[:-1].split("-")
        u1 = chooseuser(userlist=userlist, app=mo[1], platform=mo[0], accountType=mo[2], userType=mo[3],
                        changjing=mo[4])
        if u1 is None:
            continue
        content, ass_text = bulidcontent(u1, mo[4])
        if content is not None and len(content) > 0:
            # 这里在传参中加上前缀 "body=>" 以满足测试框架需要
            # line = [casename[:-1], "body=>"+json.dumps(content), ass_text]
            urlpath = "http://api2.pre.sxw.cn/passport/api/auth/login"
            if env == "online":
                urlpath = "http://api2.sxw.cn/passport/api/auth/login"
            case = caseitem(module="登录",
                            id=index_id,
                            case_name=casename[:-1],
                            method="post",
                            url=urlpath,
                            parameter=json.dumps(content),
                            assertion=ass_text,
                            isactive="y")
            line = caseitemtolist(case)
            index_id = index_id + 1
        else:
            # line = [casename[:-1], None]
            pass
        if len(line) > 1:
            ws.append(line)
    wb.save("testcase\{}\登录.xlsx".format(env))


def caseitemtolist(case: caseitem) -> List:
    """
    将caseitem转换成List类型，方便写入excel
    :param case:
    :return: List
    """
    return [case.module, case.id, case.case_name, case.method, case.url, case.parameter, case.assertion, case.isactive]


def filtercasename(casename: str) -> bool:
    """
    添加业务规则，在生成用例标题时，过滤的用例不会存在用例文件中
    :param casename: xx-xx-xx-xx-xx
    :return: <bool> True：满足过滤条件，不生成用例；False：不收过滤条件影响，生成用例
    """
    case = casename.split("-")

    # 先屏蔽龙泉登录
    if case[2] == "龙泉登录":
        return True

    # 屏蔽手机验证码登录，每天只有6次获取验证码的限制，转人工处理
    elif case[2] == "手机验证码":
        return True

    # 生学号只能在PC平台和云平台登录
    elif case[2] == "生学号":
        if case[0] == "PC" and case[1] == "云平台":
            pass
        else:
            return True

    # 生学堂类型只能通过手机登录
    elif case[1] == "生学堂":
        if case[0] == "手机":
            pass
        else:
            return True

    return False
