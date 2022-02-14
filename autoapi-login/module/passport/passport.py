# encoding = utf-8
import unittest
import gc
import time
from openpyxl import load_workbook
from common.Log import log
from domain.caseitem import caseitem
from ddt import ddt, data
from common.api import api
from BeautifulReport import BeautifulReport


def inittestcase(env: str):
    filepath = "testcase/{}/登录.xlsx".format(env)
    wb = load_workbook(filepath)
    ws = wb.worksheets[0]
    rows = ws.max_row + 1
    itemlist = []
    startindex = 1
    for i in range(2, rows):
        itemlist.append(caseitem(module=ws.cell(i, startindex).value, id=ws.cell(i, startindex + 1).value,
                                 case_name=ws.cell(i, startindex + 2).value, method=ws.cell(i, startindex + 3).value,
                                 url=ws.cell(i, startindex + 4).value, parameter=ws.cell(i, startindex + 5).value,
                                 assertion=ws.cell(i, startindex + 6).value,
                                 isactive=ws.cell(i, startindex + 7).value))
    del ws, wb
    gc.collect()
    return itemlist


@ddt
class passport_online(unittest.TestCase):
    caselist = inittestcase("online")

    @data(*caselist)
    def test_ddt_testcase(self, case: caseitem):
        if case.isactive.lower() != "y":
            self.skipTest("不执行")
        retext = api(case)
        log.error(case.ToDict())
        log.error(retext)
        self.assertIn(case.assertion, retext)
        # /api/sms/send_auth_code 短信校验

    def suite(self):
        suite = unittest.TestSuite()
        suite.addTests(unittest.TestLoader().loadTestsFromTestCase(passport_online))
        return suite

    def run_testcase(self):
        suite = self.suite()
        report_path = "report/online"
        report_file = "测试报告" + time.strftime('%Y-%m-%d-%H-%M-%S') + ".html"
        result = BeautifulReport(suite)
        result.report(filename=report_file, description="PASSPORT", report_dir=report_path, theme="theme_cyan")


@ddt
class passport_pre(unittest.TestCase):
    caselist = inittestcase("pre")

    @data(*caselist)
    def test_ddt_testcase(self, case: caseitem):
        if case.isactive.lower() != "y":
            self.skipTest("不执行")
        retext = api(case)
        log.error(case.ToDict())
        log.error(retext)
        self.assertIn(case.assertion, retext)

    def suite(self):
        suite = unittest.TestSuite()
        suite.addTests(unittest.TestLoader().loadTestsFromTestCase(passport_pre))
        return suite

    def run_testcase(self):
        suite = self.suite()
        report_path = "report/pre"
        report_file = "测试报告" + time.strftime('%Y-%m-%d-%H-%M-%S') + ".html"
        result = BeautifulReport(suite)
        result.report(filename=report_file, description="PASSPORT", report_dir=report_path, theme="theme_cyan")
